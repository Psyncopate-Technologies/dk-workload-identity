"""
Generate docs/app-registrations.xlsx — one row per (env, domain, workload).

Pairs with docs/DK-Confluent-Entra-App-Runbook.docx: each step in the runbook
points the Entra admin at a column by letter.

Usage:
    pip install openpyxl --break-system-packages   # already pinned in tools/versions.env
    python docs/generate_app_registrations_spreadsheet.py
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).parent.parent
OUT = Path(__file__).parent / "app-registrations.xlsx"

# DKP facts
TENANT_ID = "7bab0bc1-bb61-48d7-b2d4-79825c2ac6b8"

# (env, domain, workload) per Rukai's topic list. Example given: dk-confluent-dev-mergerarb-madam.
# domain/workload names to be confirmed with DKP (see CHECKLIST.md).
WORKLOADS = [
    ("mergerarb", "madam"),
    ("position",  "writer"),
    ("position",  "reader"),
    ("tft",       "writer"),
    ("ibconnect", "connector"),
]
ENVS = ["dev", "uat", "prd"]

HEADERS = [
    ("Row",                             6),
    ("Env",                             6),
    ("Domain",                          12),
    ("Workload",                        12),
    ("App Display Name",                38),
    ("Supported account types",         34),
    ("Redirect URI",                    18),
    ("Application ID URI",              38),
    ("Scope: name & value",             22),
    ("Scope: admin consent display",    26),
    ("Scope: admin consent description",44),
    ("Scope: state",                    12),
    ("Manifest: requestedAccessTokenVersion", 18),
    ("Entra Tenant ID",                 40),
    ("Application (client) ID — FILL AFTER CREATION",  40),
    ("Object ID — FILL AFTER CREATION", 40),
    ("Confluent identity pool (display name)", 38),
    ("Notes",                           40),
]

TITLE       = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="003E7E")
FILL_IN     = PatternFill("solid", fgColor="FFF2CC")
CONST       = Font(color="555555")
MONO        = Font(name="Menlo", size=9)
THIN        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
TOP         = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


def make_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "App Registrations"

    # Header row
    for col_idx, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = TITLE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "E2"

    row_num = 0
    for env in ENVS:
        for domain, workload in WORKLOADS:
            row_num += 1
            app_name = f"dk-confluent-{env}-{domain}-{workload}"
            values = [
                row_num,                                               # A Row
                env,                                                   # B Env
                domain,                                                # C Domain
                workload,                                              # D Workload
                app_name,                                              # E App display name
                "Accounts in this organizational directory only (Single tenant)",  # F
                "(leave blank — not needed for client-credentials flow)",           # G
                "api://<Application-client-ID>  (accept Azure's default after registration)",  # H
                "access_as_application",                               # I
                "access_as_application",                               # J
                f"Allow the application to access Kafka on behalf of the {domain}-{workload} workload.",  # K
                "Enabled",                                             # L
                "2",                                                   # M
                TENANT_ID,                                             # N
                "",                                                    # O fill after
                "",                                                    # P fill after
                app_name,                                              # Q Confluent pool name = app name
                ("PRD topic list not finalized — confirm with Rukai before apply." if env == "prd" else ""),
            ]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row_num + 1, column=col_idx, value=v)
                cell.border = BORDER
                cell.alignment = LEFT if col_idx >= 5 else CENTER
                # Highlight the two fields the admin fills after the portal step
                if col_idx in (15, 16):
                    cell.fill = FILL_IN
                elif col_idx in (6, 9, 10, 11, 12, 13, 14):
                    cell.font = CONST
            ws.row_dimensions[row_num + 1].height = 32

    # Legend row at the bottom
    legend_row = row_num + 3
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=legend_row + 1, column=1, value="Yellow fill = fill in after creating the app registration.")
    ws.cell(row=legend_row + 2, column=1, value="Grey text = constant value per the runbook; do not change.")
    ws.cell(row=legend_row + 3, column=1, value="Row numbers map to section 'Row N' in the runbook.")
    ws.cell(row=legend_row + 5, column=1, value="After every row is complete, email the spreadsheet to the Confluent PS team.")

    return wb


def load_workloads_json(env: str) -> dict:
    path = REPO_ROOT / "terraform" / "live" / env / "workloads.json"
    return json.loads(path.read_text())


def fmt_list(label: str, items: list[str]) -> str:
    if not items:
        return ""
    return f"{label}:\n" + "\n".join(f"  • {x}" for x in items)


def access_cell(wl: dict, prefix_field: str, name_field: str) -> str:
    parts = []
    prefs = wl.get(prefix_field, [])
    names = wl.get(name_field, [])
    if prefs:
        parts.append(fmt_list("prefix", prefs))
    if names:
        parts.append(fmt_list("exact",  names))
    return "\n\n".join(parts)


def add_rbac_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("App ↔ Pool ↔ RBAC")

    headers = [
        ("Row",                           6),
        ("Env",                           6),
        ("Domain",                        12),
        ("Workload",                      12),
        ("Entra App Registration",        34),
        ("Confluent Identity Pool",       34),
        ("Pool filter (templated)",       50),
        ("Write topics",                  40),
        ("Read topics",                   40),
        ("Manage topics (DeveloperManage)", 40),
        ("Consumer groups (DeveloperRead)", 36),
        ("Kafka cluster",                 16),
    ]
    for col_idx, (name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = TITLE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "E2"

    row_num = 0
    for env in ENVS:
        cfg = load_workloads_json(env)
        cluster = cfg.get("kafka_cluster_id", "?")
        tenant  = cfg.get("entra_tenant_id", "?")

        for (domain, workload) in WORKLOADS:
            row_num += 1
            key = f"{domain}-{workload}"
            wl = cfg.get("workloads", {}).get(key, {})
            app_name = f"dk-confluent-{env}-{domain}-{workload}"

            pool_filter = (
                f'claims.tid == "{tenant}" && '
                f'claims.aud == "<Application (client) ID GUID>"'
            )

            values = [
                row_num,
                env,
                domain,
                workload,
                app_name,
                app_name,  # pool = app name
                pool_filter,
                access_cell(wl, "write_topic_prefixes", "write_topic_names"),
                access_cell(wl, "read_topic_prefixes",  "read_topic_names"),
                access_cell(wl, "manage_topic_prefixes", "manage_topic_names"),
                access_cell(wl, "consumer_group_prefixes", "consumer_group_names"),
                cluster,
            ]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row_num + 1, column=col_idx, value=v)
                cell.border = BORDER
                if col_idx in (8, 9, 10, 11):
                    cell.alignment = TOP
                    cell.font = MONO
                elif col_idx == 7:
                    cell.alignment = TOP
                    cell.font = MONO
                elif col_idx >= 5:
                    cell.alignment = LEFT
                else:
                    cell.alignment = CENTER
            # Dynamic row height based on max lines in the access cells
            max_lines = max(
                (str(v).count("\n") + 1 if v else 1)
                for v in values[6:11]
            )
            ws.row_dimensions[row_num + 1].height = max(24, min(max_lines * 13, 260))

    # Legend
    legend_row = row_num + 3
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    ws.cell(row=legend_row + 1, column=1,
            value="prefix: CRN topic=<value>*  (matches all topics beginning with <value>)")
    ws.cell(row=legend_row + 2, column=1,
            value="exact:  CRN topic=<value>   (matches that topic only)")
    ws.cell(row=legend_row + 3, column=1,
            value="Pool filter's aud is the Application (client) ID GUID (v2 Entra token shape) — no api:// prefix.")
    ws.cell(row=legend_row + 5, column=1,
            value="Source of truth: terraform/live/<env>/workloads.json — re-run this generator after edits.")


if __name__ == "__main__":
    wb = make_workbook()
    add_rbac_sheet(wb)
    wb.save(OUT)
    print(f"wrote {OUT}")
